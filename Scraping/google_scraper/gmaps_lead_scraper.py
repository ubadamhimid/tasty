import asyncio
import os
import re
import sys
import io
import sqlite3
import urllib.parse
from datetime import datetime
import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright

# Force UTF-8 encoding for standard output streams to handle Arabic text on Windows terminals
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# Configuration Constants
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.path.join(SCRIPT_DIR, "leads_history.db")

# Query Matrix Variables
CITIES = ["دمشق", "حلب", "حمص", "اللاذقية", "طرطوس", "حماة", "السويداء", "درعا"]
CATEGORIES = ["عيادة أسنان", "مدينة ملاهي", "ملعب بادل", "شركة تنظيف", "مجمع سينما"]

# ==========================================
# DATABASE HELPER METHODS (DEDUPLICATION)
# ==========================================
def init_db():
    """Initializes the SQLite database schema for lead history tracking."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT,
            website TEXT,
            maps_url TEXT UNIQUE,
            category TEXT,
            rating TEXT,
            reviews_count TEXT,
            facebook TEXT,
            instagram TEXT,
            scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def is_duplicate(maps_url, phone):
    """
    Checks if a business has already been scraped in previous runs.
    Checks unique Maps URL or phone number.
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Check by URL
    cursor.execute("SELECT 1 FROM leads WHERE maps_url = ?", (maps_url,))
    url_exists = cursor.fetchone() is not None
    
    # Check by phone (if valid and not N/A)
    phone_exists = False
    if phone and phone != "N/A":
        cursor.execute("SELECT 1 FROM leads WHERE phone = ?", (phone,))
        phone_exists = cursor.fetchone() is not None
        
    conn.close()
    return url_exists or phone_exists

def save_lead_to_db(lead):
    """Inserts a newly scraped lead record into the SQLite database history."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO leads (name, phone, website, maps_url, category, rating, reviews_count, facebook, instagram)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            lead["Business Name (اسم العمل)"],
            lead["Phone Number (رقم الهاتف)"],
            lead["Website URL (رابط الموقع)"],
            lead["Maps URL (رابط الخريطة)"],
            lead["Business Category (التصنيف)"],
            lead["Rating (التقييم)"],
            lead["Reviews Count (عدد المراجعات)"],
            lead["Facebook Link (فيسبوك)"],
            lead["Instagram Link (إنستغرام)"]
        ))
        conn.commit()
    except sqlite3.IntegrityError:
        pass # Handle duplicate collision gracefully
    finally:
        conn.close()

async def process_single_listing(sem, context, listing_url, idx, total_count, query):
    """
    Processes a single Maps listing URL inside a concurrent tab.
    Extracts name, phone, category, website, ratings, checks socials,
    runs DDG fallbacks and Facebook phone scraper fallback if needed.
    """
    async with sem:
        page = await context.new_page()
        try:
            # Navigate to the place details page (using domcontentloaded for performance)
            await page.goto(listing_url, wait_until="domcontentloaded")
            try:
                await page.wait_for_selector('h1', timeout=10000)
            except Exception:
                pass
            await page.wait_for_timeout(800) # Small delay to let page stabilize
            
            # Extract Business Name
            name = ""
            name_selectors = [
                'h1.DUwDvf', 
                'h1.fontHeadlineLarge',
                'div.lMbQif h1',
                'h1'
            ]
            for sel in name_selectors:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    name = await loc.first.text_content()
                    name = name.strip() if name else ""
                    break
            
            if not name:
                return None
                
            # Extract Phone Number
            phone = "N/A"
            try:
                phone_selectors = [
                    'a[href^="tel:"]',
                    'button[data-item-id^="phone:tel:"]',
                    'button[data-value^="tel:"]',
                    'a[data-item-id^="phone:tel:"]'
                ]
                for p_sel in phone_selectors:
                    loc = page.locator(p_sel)
                    if await loc.count() > 0:
                        href = await loc.first.get_attribute("href") or await loc.first.get_attribute("data-item-id") or await loc.first.get_attribute("data-value")
                        if href:
                            phone = href.replace("tel:", "").replace("phone:", "").strip()
                            break
            except Exception:
                pass
                
            # Deduplicate by phone immediately if found on maps
            if phone != "N/A" and is_duplicate(None, phone):
                print(f"      [{idx}/{total_count}] Skipping (Duplicate Phone Number: {phone})")
                return None
                
            # Extract Rating and Reviews Count
            rating = "N/A"
            reviews_count = "0"
            try:
                rating_loc = page.locator('div.F7nice span[aria-hidden="true"]').first
                if await rating_loc.count() > 0:
                    rating = await rating_loc.text_content()
                    rating = rating.strip() if rating else "N/A"
                
                reviews_loc = page.locator('div.F7nice span[aria-label*="مراجعة"], div.F7nice span[aria-label*="review"]').first
                if await reviews_loc.count() > 0:
                    reviews_text = await reviews_loc.get_attribute("aria-label")
                    reviews_match = re.search(r'\d+', reviews_text)
                    if reviews_match:
                        reviews_count = reviews_match.group()
            except Exception:
                pass
                
            # Extract Business Category
            category = "N/A"
            try:
                cat_selectors = [
                    'button[jsaction*="pane.rating.category"]',
                    'button[jsaction*="category"]',
                    'span.fontBodyMedium button',
                    'span[class*="fontBodyMedium"] button'
                ]
                for c_sel in cat_selectors:
                    loc = page.locator(c_sel)
                    if await loc.count() > 0:
                        cat_text = await loc.first.text_content()
                        if cat_text:
                            category = cat_text.strip()
                            break
            except Exception:
                pass
                
            # Extract Website URL
            website = "N/A"
            try:
                web_selectors = [
                    'a[data-item-id="authority"]',
                    'a[data-tooltip*="website"]',
                    'a[data-tooltip*="الموقع الإلكتروني"]',
                    'a[aria-label*="الموقع الإلكتروني"]',
                    'a[aria-label*="website"]'
                ]
                for w_sel in web_selectors:
                    loc = page.locator(w_sel)
                    if await loc.count() > 0:
                        web_href = await loc.first.get_attribute("href")
                        if web_href and "google.com" not in web_href:
                            website = web_href.strip()
                            break
            except Exception:
                pass
                
            if website and website != "N/A":
                website_clean = website.split('?')[0].split('#')[0]
            else:
                website_clean = "N/A"
                
            # Website evaluation
            is_custom_website = False
            has_free_site = False
            if website_clean != "N/A":
                if "business.site" in website_clean or "google.com" in website_clean:
                    has_free_site = True
                else:
                    is_custom_website = True
            
            # Extract Social Media Profiles
            facebook = "N/A"
            instagram = "N/A"
            try:
                social_links = await page.locator('a[href*="facebook.com"], a[href*="instagram.com"]').all()
                for link in social_links:
                    href = await link.get_attribute("href")
                    if href:
                        if "facebook.com" in href and facebook == "N/A":
                            if not any(x in href for x in ["sharer", "login", "pages/create", "help"]):
                                facebook = href
                        elif "instagram.com" in href and instagram == "N/A":
                            if not any(x in href for x in ["developer", "about", "press"]):
                                instagram = href
            except Exception:
                pass
                
            # DDG Fallback Lookup if website or socials missing (using same page instance)
            if website_clean == "N/A" and (facebook == "N/A" or instagram == "N/A"):
                try:
                    search_query = f"{name} {query.replace('في', '').strip()} facebook instagram"
                    ddg_url = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote(search_query)}"
                    await page.goto(ddg_url, wait_until="domcontentloaded")
                    await page.wait_for_timeout(600)
                    
                    results_links = await page.locator('a.result__url').all()
                    for r_link in results_links:
                        href = await r_link.get_attribute("href")
                        if href:
                            decoded_href = urllib.parse.unquote(href)
                            match = re.search(r'uddg=(https?://[^&]+)', decoded_href)
                            real_url = match.group(1) if match else decoded_href
                            
                            if "facebook.com" in real_url and facebook == "N/A":
                                if not any(x in real_url for x in ["sharer", "login", "pages/create", "help"]):
                                    facebook = real_url
                            elif "instagram.com" in real_url and instagram == "N/A":
                                if not any(x in real_url for x in ["developer", "about", "press"]):
                                    instagram = real_url
                                    
                        if facebook != "N/A" and instagram != "N/A":
                            break
                except Exception:
                    pass
            
            # Facebook Phone Scraper Fallback (using same page instance)
            if phone == "N/A" and facebook != "N/A":
                try:
                    await page.goto(facebook, wait_until="domcontentloaded", timeout=12000)
                    await page.wait_for_timeout(1000)
                    
                    tel_loc = page.locator('a[href^="tel:"]')
                    if await tel_loc.count() > 0:
                        href = await tel_loc.first.get_attribute("href")
                        if href:
                            phone = href.replace("tel:", "").strip()
                    else:
                        content = await page.content()
                        phone_matches = re.findall(r'\+963\s?\d{2,3}\s?\d{3,4}\s?\d{3,4}|\b09\d{8}\b|\b011\d{6,7}\b', content)
                        if phone_matches:
                            phone = phone_matches[0].strip()
                except Exception:
                    pass
            
            # Double-check phone duplicate history
            if phone != "N/A" and is_duplicate(None, phone):
                print(f"      [{idx}/{total_count}] Skipping (Duplicate Phone Number: {phone})")
                return None
            
            # Success log print details
            print(f"      [✔] Extracted: {name} | Category: {category} | Phone: {phone} | Web: {website_clean}")
            
            lead_data = {
                "Query": query,
                "Business Name (اسم العمل)": name,
                "Business Category (التصنيف)": category,
                "Phone Number (رقم الهاتف)": phone,
                "Website URL (رابط الموقع)": website_clean,
                "Is Custom Website (موقع مخصص)": "Yes" if is_custom_website else "No",
                "Free Business.site (موقع مجاني)": "Yes" if has_free_site else "No",
                "Facebook Link (فيسبوك)": facebook,
                "Instagram Link (إنستغرام)": instagram,
                "Rating (التقييم)": rating,
                "Reviews Count (عدد المراجعات)": reviews_count,
                "Maps URL (رابط الخريطة)": listing_url
            }
            
            save_lead_to_db(lead_data)
            return lead_data
            
        except Exception as err:
            print(f"      [-] Error extracting details for {listing_url}: {err}")
            return None
        finally:
            await page.close()

# ==========================================
# SCRAPING ENGINE
# ==========================================
async def scrape_google_maps(queries, max_leads_per_query, headed=False):
    """
    Asynchronously runs searches on Google Maps, scrolls through feed panels,
    extracts business attributes, performs fallback social lookups, and applies deduplication filters.
    """
    init_db()
    results = []

    async with async_playwright() as p:
        # Launch Chromium headless or headed depending on background environment
        browser = await p.chromium.launch(
            headless=not headed,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-infobars"
            ]
        )
        
        # Configure localized context
        context = await browser.new_context(
            user_agent=USER_AGENT,
            viewport={"width": 1280, "height": 800},
            locale="ar"
        )
        
        # Stealth logic to bypass bot detection signatures
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
        """)
        
        page = await context.new_page()
        
        for q_idx, query in enumerate(queries, 1):
            print(f"\n[*] [{q_idx}/{len(queries)}] Starting search for: '{query}'")
            try:
                # Load Maps homepage
                await page.goto("https://www.google.com/maps", wait_until="domcontentloaded")
                await page.wait_for_timeout(2000)
                
                # Robust cookie consent handler
                try:
                    consent_selectors = [
                        'button[aria-label="Accept all"]',
                        'button:has-text("Accept all")',
                        'button:has-text("الموافقة على الكل")',
                        'button:has-text("موافق")',
                        'form[action*="consent.google.com"] button',
                        'button[aria-label*="الموافقة"]'
                    ]
                    for sel in consent_selectors:
                        btn = page.locator(sel)
                        if await btn.count() > 0:
                            print(f"    [Info] Found and clicking cookie consent button: {sel}")
                            await btn.first.click()
                            
                            # Wait for page redirect to finish
                            try:
                                await page.wait_for_function('() => window.location.href.includes("/maps") || !!document.querySelector("input#searchboxinput")', timeout=15000)
                            except Exception:
                                pass
                            await page.wait_for_timeout(2000)
                            break
                except Exception as e:
                    print(f"    [Info] Consent dialog check failed: {e}")
                
                # Re-verify redirect out of consent pages
                try:
                    await page.wait_for_selector('input#searchboxinput', timeout=15000)
                except Exception:
                    if "consent.google.com" in page.url or "consent" in page.url:
                        print("    [Warning] Still on consent page, attempting to re-navigate to Maps...")
                        await page.goto("https://www.google.com/maps", wait_until="domcontentloaded")
                        try:
                            await page.wait_for_selector('input#searchboxinput', timeout=15000)
                        except Exception:
                            pass
                        await page.wait_for_timeout(3000)
                
                # Locate search box, focus, and execute search query
                try:
                    search_box_selectors = [
                        'input#searchboxinput',
                        'input.searchboxinput',
                        '#searchboxinput',
                        'input[name="q"]',
                        'input[aria-label*="Search"]',
                        'input[aria-label*="البحث"]'
                    ]
                    search_box = None
                    for sel in search_box_selectors:
                        loc = page.locator(sel)
                        if await loc.count() > 0:
                            search_box = loc.first
                            break
                            
                    if not search_box:
                        raise Exception("Search box element not found in DOM with any known selectors.")
                        
                    await search_box.wait_for(state="visible", timeout=10000)
                    await search_box.click()
                    await page.wait_for_timeout(500)
                    await search_box.fill(query)
                    await page.wait_for_timeout(500)
                    await search_box.press("Enter")
                except Exception as search_err:
                    screenshot_name = f"search_error_{re.sub(r'[^a-zA-Z0-9]', '_', query)}.png"
                    screenshot_path = os.path.join(SCRIPT_DIR, screenshot_name)
                    await page.screenshot(path=screenshot_path)
                    print(f"    [Warning] Failed to find or interact with search box. Screenshot saved to {screenshot_path}")
                    raise search_err
                
                # Wait for feed results list to render
                print("    [*] Waiting for search results to load...")
                try:
                    await page.wait_for_selector('div[role="feed"]', timeout=15000)
                except Exception:
                    try:
                        await page.wait_for_selector('a[href*="/maps/place/"]', timeout=10000)
                    except Exception as feed_err:
                        screenshot_name = f"feed_error_{re.sub(r'[^a-zA-Z0-9]', '_', query)}.png"
                        screenshot_path = os.path.join(SCRIPT_DIR, screenshot_name)
                        await page.screenshot(path=screenshot_path)
                        print(f"    [Warning] Failed to find listings feed. Screenshot saved to {screenshot_path}")
                        raise feed_err
                
                # Scroll results panel to fetch matching URLs
                feed_selector = 'div[role="feed"]'
                feed_exists = await page.locator(feed_selector).count() > 0
                
                if feed_exists:
                    feed = page.locator(feed_selector).first
                    print("    [*] Scrolling left panel to load listings...")
                    
                    listings_set = set()
                    previous_count = 0
                    no_new_results_count = 0
                    
                    while len(listings_set) < max_leads_per_query:
                        # Extract listing links via robust page evaluate (prevents DOM detachment errors)
                        hrefs = await page.evaluate('''() => {
                            const links = Array.from(document.querySelectorAll('div[role="feed"] a[href*="/maps/place/"]'));
                            return links.map(a => a.href);
                        }''')
                        for href in hrefs:
                            listings_set.add(href)
                        
                        current_count = len(listings_set)
                        
                        if current_count >= max_leads_per_query:
                            break
                            
                        if current_count == previous_count:
                            no_new_results_count += 1
                            if no_new_results_count > 20:
                                break
                        else:
                            no_new_results_count = 0
                            
                        previous_count = current_count
                        
                        # Scroll down the feed container directly in JS
                        await page.evaluate('''() => {
                            const feedEl = document.querySelector('div[role="feed"]');
                            if (feedEl) {
                                feedEl.scrollTop = feedEl.scrollHeight;
                            }
                        }''')
                        await page.wait_for_timeout(1500)
                else:
                    # Direct page redirect when only a single result is found
                    current_url = page.url
                    if "/maps/place/" in current_url:
                        listings_set = {current_url}
                    else:
                        print("    [-] Could not locate listings results feed.")
                        continue
                
                # Filter leads and process details concurrently
                listings_to_scrape = list(listings_set)[:max_leads_per_query]
                print(f"    [*] Found {len(listings_to_scrape)} leads. Processing details concurrently (Max 5 parallel workers)...")
                
                sem = asyncio.Semaphore(5)  # Restrict to 5 concurrent browser tabs
                tasks = []
                for idx, listing_url in enumerate(listings_to_scrape, 1):
                    # Check database history before initiating worker task
                    if is_duplicate(listing_url, None):
                        print(f"      [{idx}/{len(listings_to_scrape)}] Skipping (Duplicate URL): {listing_url.split('/place/')[1][:30]}...")
                        continue
                        
                    tasks.append(process_single_listing(sem, context, listing_url, idx, len(listings_to_scrape), query))
                
                # Execute concurrently
                scraped_leads = await asyncio.gather(*tasks)
                
                # Record successful leads
                for lead in scraped_leads:
                    if lead:
                        results.append(lead)
                        
            except Exception as query_err:
                print(f"[-] Error processing query '{query}': {query_err}")
                
        await browser.close()
        
    return results

# ==========================================
# OUTPUT GENERATOR (STYLED EXCEL WORKBOOK)
# ==========================================
def format_excel_output(df, file_path):
    """
    Saves and formats data into a professionally styled Excel workbook.
    Highlights hot prospects (no custom website) in a soft yellow color fill.
    If file exists, appends data safely and applies styling over the entire dataset.
    """
    # If Excel file already exists, load and append data, then deduplicate
    if os.path.exists(file_path):
        try:
            existing_df = pd.read_excel(file_path)
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            # Remove duplicates by Unique Maps URL
            combined_df.drop_duplicates(subset=["Maps URL (رابط الخريطة)"], keep="first", inplace=True)
            df = combined_df
            print(f"[*] Appending to existing Excel file. Total records in sheet: {len(df)}")
        except Exception as e:
            print(f"[-] Warning: Failed to read existing Excel file: {e}. Rewriting file.")

    # Write dataframe to Excel worksheet (with retry handler if locked by Excel or another process)
    import time
    writer = None
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        try:
            writer = pd.ExcelWriter(file_path, engine="openpyxl")
            break
        except PermissionError:
            if attempt == max_retries:
                # Fallback to a modified timestamped filename if permanently locked
                timestamp = datetime.now().strftime("%H%M%S")
                base, ext = os.path.splitext(file_path)
                file_path = f"{base}_{timestamp}{ext}"
                print(f"[Warning] Excel file is permanently locked. Saving output to alternative file path: {file_path}")
                writer = pd.ExcelWriter(file_path, engine="openpyxl")
            else:
                print(f"[Warning] Excel file is locked. Retrying in 2 seconds... (Attempt {attempt}/{max_retries})")
                time.sleep(2)
                
    df.to_excel(writer, sheet_name="Leads Output", index=False)
    
    workbook = writer.book
    worksheet = writer.sheets["Leads Output"]
    
    # Enable grid lines visibility
    worksheet.views.sheetView[0].showGridLines = True
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Palette definition
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Corporate Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid") # Slate zebra tint
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    flag_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Warm Yellow highlighting for prime target leads
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Headers Styling
    for col_idx, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    worksheet.row_dimensions[1].height = 28
    
    # Rows Styling & Highlighting
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 22
        
        # Identify if lead lacks a custom website
        is_custom_web = worksheet.cell(row=row_idx, column=df.columns.get_loc("Is Custom Website (موقع مخصص)") + 1).value
        
        # Apply zebra coloring or yellow target highlight
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        if is_custom_web == "No":
            row_fill = flag_fill
            
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            
            # Align numeric scores, rating, phones, and flags to center
            col_name = df.columns[col_idx - 1]
            if any(term in col_name for term in ["Phone", "Rating", "Count", "Website", "Link", "Is Custom", "Free"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    # Auto-adjust column widths dynamically with safety padding
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val) + 4)
            else:
                max_len = max(max_len, len(val))
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    writer.close()
    print(f"\n[+] Lead output saved and formatted successfully in: {file_path}")

# ==========================================
# MAIN EXECUTION ROUTINE
# ==========================================
async def main():
    print("=" * 60)
    print("        ENTERPRISE GOOGLE MAPS LEAD GENERATION ENGINE")
    print("=" * 60)
    
    # Programmatically compile Query Matrix (Categories × Cities)
    queries = [f"{cat} في {city}" for cat in CATEGORIES for city in CITIES]
    
    # Generate timestamped filename
    today_str = datetime.now().strftime("%Y_%m_%d")
    output_filename = f"leads_output_{today_str}.xlsx"
    output_path = os.path.join(SCRIPT_DIR, output_filename)
    
    print(f"[*] Dynamically generated {len(queries)} query combinations.")
    print(f"[*] Output destination: {output_path}")
    print(f"[*] Deduplication History: {DB_NAME}")
    print("=" * 60)
    
    max_leads_per_query = 50  # Upgraded target maximum leads per query
    
    print(f"[*] Running search on full query matrix ({len(queries)} combinations)...")
    print(f"[*] Target max leads per query: {max_leads_per_query}")
    print("-" * 60)
    
    # Initialize and execute scraping engine
    results = await scrape_google_maps(queries, max_leads_per_query, headed=False)
    
    if results:
        # Convert to DataFrame
        df = pd.DataFrame(results)
        # Save and style Excel worksheet
        format_excel_output(df, output_path)
    else:
        print("\n[-] No new leads were extracted during this run (all matches were filtered out as duplicates).")

if __name__ == "__main__":
    asyncio.run(main())
