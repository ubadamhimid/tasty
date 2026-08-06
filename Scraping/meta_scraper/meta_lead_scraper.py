import asyncio
import os
import re
import sys
import random
import sqlite3
import urllib.parse
from datetime import datetime
import pandas as pd
from playwright.async_api import async_playwright

# Force UTF-8 encoding for standard output streams to handle Arabic text on Windows terminals
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

# Configuration Constants
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = os.path.join(SCRIPT_DIR, "meta_leads_history.db")
USER_DATA_DIR = os.path.join(SCRIPT_DIR, "meta_user_data")
LOGIN_MARKER_PATH = os.path.join(SCRIPT_DIR, "login_success.txt")

# Query Matrix Variables
CITIES = ["دمشق", "حلب", "حمص", "اللاذقية", "طرطوس", "حماة", "السويداء", "درعا"]
CATEGORIES = ["عيادة أسنان", "مدينة ملاهي", "ملعب بادل", "شركة تنظيف", "مجمع سينما"]

# ==========================================
# DATABASE HELPER METHODS (DEDUPLICATION)
# ==========================================
def init_db():
    """Initializes the SQLite database schema for Facebook leads history tracking."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            phone TEXT,
            email TEXT,
            website TEXT,
            facebook_url TEXT UNIQUE,
            category TEXT,
            scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    conn.close()

def is_duplicate(facebook_url, phone):
    """
    Checks if a Facebook Page URL or Phone Number has already been scraped.
    """
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Check by Facebook URL
    url_exists = False
    if facebook_url:
        cursor.execute("SELECT 1 FROM leads WHERE facebook_url = ?", (facebook_url,))
        url_exists = cursor.fetchone() is not None
    
    # Check by phone (if valid and not N/A)
    phone_exists = False
    if phone and phone != "N/A":
        cursor.execute("SELECT 1 FROM leads WHERE phone = ?", (phone,))
        phone_exists = cursor.fetchone() is not None
        
    conn.close()
    return url_exists or phone_exists

def save_lead_to_db(lead):
    """Inserts a newly scraped Facebook lead record into the database."""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO leads (name, phone, email, website, facebook_url, category)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            lead["Business Name (اسم العمل)"],
            lead["Phone Number (رقم الهاتف)"],
            lead["Email (البريد الإلكتروني)"],
            lead["Website URL (رابط الموقع)"],
            lead["Facebook Link (رابط فيسبوك)"],
            lead["Business Category (التصنيف)"]
        ))
        conn.commit()
    except sqlite3.IntegrityError:
        pass  # Handle duplicate collision gracefully
    finally:
        conn.close()

# ==========================================
# URL CLEANING & NORMALIZATION
# ==========================================
def clean_facebook_url(url):
    """
    Cleans and standardizes Facebook URLs to isolate page/profile roots.
    Returns standard facebook.com/username or facebook.com/pages/name/id.
    """
    if not url:
        return None
    
    # Normalize relative links
    if url.startswith('/'):
        url = "https://www.facebook.com" + url
        
    parsed = urllib.parse.urlparse(url)
    if "facebook.com" not in parsed.netloc:
        return None
        
    path = parsed.path
    path_parts = [p for p in path.split('/') if p]
    if not path_parts:
        return None
        
    first_part = path_parts[0]
    
    # System exclusions we shouldn't scrape
    exclusions = {
        "groups", "events", "messages", "notifications", "friends", "marketplace", 
        "gaming", "watch", "search", "policies", "about", "help", "privacy", 
        "legal", "login", "reg", "recover", "stories", "sharer", 
        "ajax", "photo", "video", "permalink", "posts", "ads", "business", "campaign",
        "home", "settings", "notifications", "r.php", "checkpoint", "create", "home.php",
        "intern", "jobs", "local", "memories", "menus", "news", "safety", "services",
        "offers", "places", "saved", "pages_manager", "live", "reels", "profile"
    }
    
    # Check special profile/page patterns
    if first_part == "pages" and len(path_parts) >= 2:
        clean_path = "/".join(path_parts[:3])
        return f"https://www.facebook.com/{clean_path}"
        
    if first_part == "people" and len(path_parts) >= 2:
        clean_path = "/".join(path_parts[:3])
        return f"https://www.facebook.com/{clean_path}"
        
    if first_part == "profile.php":
        qs = urllib.parse.parse_qs(parsed.query)
        if "id" in qs:
            return f"https://www.facebook.com/profile.php?id={qs['id'][0]}"
        return None
        
    if first_part in exclusions:
        return None
        
    # Standard profile/page username
    return f"https://www.facebook.com/{first_part}"

def get_about_url(url):
    """
    Constructs the correct /about sub-URL for standard pages and profile.php links.
    """
    parsed = urllib.parse.urlparse(url)
    if "profile.php" in parsed.path:
        qs = urllib.parse.parse_qs(parsed.query)
        if "id" in qs:
            return f"https://www.facebook.com/profile.php?id={qs['id'][0]}&sk=about"
    base_url = url.rstrip('/')
    return f"{base_url}/about"

# ==========================================
# HUMAN BEHAVIOR SIMULATORS
# ==========================================
async def random_delay():
    """Adds a human-like delay between page visits/requests."""
    delay = random.uniform(4.0, 9.0)
    print(f"      [Info] Waiting {delay:.2f} seconds to mimic human activity...")
    await asyncio.sleep(delay)

async def human_scroll(page, distance_total=1200):
    """Scrolls pages smoothly in randomized intervals."""
    scroll_chunk = 200
    current_scroll = 0
    while current_scroll < distance_total:
        chunk = random.randint(scroll_chunk - 50, scroll_chunk + 50)
        await page.evaluate(f"window.scrollBy(0, {chunk})")
        current_scroll += chunk
        await asyncio.sleep(random.uniform(0.3, 0.7))

async def human_mouse_move(page):
    """Generates random mouse movements on the screen to trigger UI renders and avoid bot flags."""
    try:
        for _ in range(random.randint(1, 3)):
            x = random.randint(150, 900)
            y = random.randint(150, 600)
            steps = random.randint(10, 20)
            await page.mouse.move(x, y, steps=steps)
            await asyncio.sleep(random.uniform(0.2, 0.4))
    except Exception:
        pass

# ==========================================
# AUTHENTICATION SESSION HANDLER
# ==========================================
async def launch_browser_context(playwright, user_data_dir, headless):
    """
    Launches a persistent browser context. Attempts to use Google Chrome (stable channel)
    for premium bot avoidance, and falls back to standard Chromium if Chrome is not installed.
    """
    try:
        context = await playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=headless,
            channel="chrome",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-infobars"
            ],
            user_agent=USER_AGENT,
            viewport={"width": 1280, "height": 800},
            locale="ar"
        )
        print("[*] Successfully launched persistent context using Google Chrome.")
        return context
    except Exception as e:
        print(f"[Info] Google Chrome launch failed ({e}). Falling back to standard Playwright Chromium...")
        context = await playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-infobars"
            ],
            user_agent=USER_AGENT,
            viewport={"width": 1280, "height": 800},
            locale="ar"
        )
        print("[*] Launched persistent context using standard Playwright Chromium.")
        return context

async def setup_auth_context(playwright, headed=False):
    """
    Initializes persistent browser session. If a login success marker doesn't exist,
    boots Chrome in headed mode, pauses for 60s for manual login and captcha resolution,
    then creates a marker. Otherwise, runs in headless mode.
    """
    if not os.path.exists(LOGIN_MARKER_PATH):
        print("\n" + "!" * 70)
        print("[!] LOGIN SESSION NOT CONFIGURED!")
        print(f"[!] Path: {USER_DATA_DIR}")
        print("[!] Launching Chrome/Chromium in HEADED mode for manual login.")
        print("[!] You have 60 seconds to log in to Facebook.")
        print("[!] Solve any captchas, robot validations, or checkpoints that appear.")
        print("[!] Once logged in, the session is saved permanently in your local user data directory.")
        print("!" * 70 + "\n")
        
        context = await launch_browser_context(playwright, USER_DATA_DIR, headless=False)
        
        for page in context.pages:
            await page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
            """)
            
        page = context.pages[0] if context.pages else await context.new_page()
        await page.goto("https://www.facebook.com/", wait_until="domcontentloaded")
        
        # Wait for user to confirm successful login
        print("\n" + "*" * 70)
        print("[!] ACTION REQUIRED:")
        print("[!] 1. Go to the browser window that just opened.")
        print("[!] 2. Log in to your Facebook account.")
        print("[!] 3. Solve any checkpoints, robot checks (captcha), or security questions.")
        print("[!] 4. Once you are logged in and see your Facebook homepage/newsfeed, return to this terminal.")
        print("[!] 5. Press [ENTER] in this terminal to save the session and start scraping.")
        print("*" * 70 + "\n")
        
        await asyncio.get_event_loop().run_in_executor(
            None,
            input,
            "--> Press [ENTER] here once you have logged in successfully: "
        )
        
        print("[*] Saving persistent profile and creating login success marker...")
        with open(LOGIN_MARKER_PATH, "w", encoding="utf-8") as f:
            f.write(f"Logged in at {datetime.now().isoformat()}")
        print("[+] Persistent login state saved in: " + USER_DATA_DIR)
        
        return context, True
    else:
        print(f"[*] Loading persistent Facebook profile from: {USER_DATA_DIR}")
        context = await launch_browser_context(playwright, USER_DATA_DIR, headless=not headed)
        
        for page in context.pages:
            await page.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
            """)
        return context, False

# ==========================================
# PAGE DETAILS EXTRACTOR
# ==========================================
async def extract_details_from_page(page, page_url):
    """
    Parses current page DOM to extract Name, Category, Phone, Email, and Custom Website.
    """
    # 1. Extract Page Name (H1)
    name = "N/A"
    try:
        loc = page.locator("h1")
        if await loc.count() > 0:
            name_text = await loc.first.text_content()
            if name_text:
                name = name_text.strip()
    except Exception:
        pass
        
    if not name or name == "N/A":
        try:
            title = await page.title()
            if title:
                name = title.split('|')[0].split('-')[0].strip()
        except Exception:
            pass
            
    if not name:
        name = "N/A"
        
    # 2. Extract Category
    category = "N/A"
    try:
        content = await page.content()
        # Regex matching Facebook's 'Page · [Category]' or 'صفحة · [Category]'
        cat_match = re.search(r'(?:Page|صفحة)\s?·\s?([^\n\r<|•]+)', content)
        if cat_match:
            category = cat_match.group(1).strip()
    except Exception:
        pass
        
    # 3. Extract Phone Number
    phone = "N/A"
    try:
        tel_loc = page.locator('a[href^="tel:"]')
        if await tel_loc.count() > 0:
            href = await tel_loc.first.get_attribute("href")
            if href:
                phone = href.replace("tel:", "").strip()
        else:
            # Pattern matching Syrian landline/mobile formats or general international numbers
            text = await page.locator('body').text_content()
            phone_match = re.search(r'\+963[\s\-]?\d{2,3}[\s\-]?\d{3,4}[\s\-]?\d{3,4}|\b09\d{8}\b|\b0[1-9]\d{1,2}\d{6,7}\b', text)
            if phone_match:
                phone = phone_match.group(0).strip()
    except Exception:
        pass
        
    # 4. Extract Email Address
    email = "N/A"
    try:
        mail_loc = page.locator('a[href^="mailto:"]')
        if await mail_loc.count() > 0:
            href = await mail_loc.first.get_attribute("href")
            if href:
                email = href.replace("mailto:", "").strip()
        else:
            text = await page.locator('body').text_content()
            email_match = re.search(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
            if email_match:
                email = email_match.group(0).strip()
    except Exception:
        pass
        
    # 5. Extract Custom Website Link (l.facebook.com redirect parser)
    website = "N/A"
    try:
        web_links = await page.locator('a[href*="l.facebook.com/l.php"]').all()
        for web in web_links:
            href = await web.get_attribute("href")
            if href:
                parsed_href = urllib.parse.urlparse(href)
                qs = urllib.parse.parse_qs(parsed_href.query)
                if "u" in qs:
                    raw_url = qs["u"][0]
                    # Exclude major social domains to isolate business custom websites
                    if not any(x in raw_url.lower() for x in ["facebook.com", "instagram.com", "twitter.com", "youtube.com", "linkedin.com", "whatsapp.com", "t.me"]):
                        website = raw_url.split('?')[0].split('#')[0].strip()
                        break
    except Exception:
        pass
        
    return {
        "name": name,
        "category": category,
        "phone": phone,
        "email": email,
        "website": website
    }

async def scrape_single_page(page, page_url):
    """
    Navigates to a Facebook page, scrolls, crawls its main feed,
    and falls back to `/about` if details are incomplete.
    """
    print(f"    [*] Navigating to: {page_url}")
    try:
        await page.goto(page_url, wait_until="domcontentloaded", timeout=25000)
        await page.wait_for_timeout(2000)
        await human_mouse_move(page)
        await human_scroll(page, distance_total=800)
        
        # Verify if cookie expired / redirect to login page happened
        if "login" in page.url or "checkpoint" in page.url:
            print("      [Warning] Session expired. Facebook redirected to login page.")
            return None
            
        data = await extract_details_from_page(page, page_url)
        
        # If details are missing, run fallback crawl on the Page's About tab
        if data["phone"] == "N/A" or data["email"] == "N/A" or data["website"] == "N/A":
            about_url = get_about_url(page_url)
            print(f"      [Info] Details incomplete. Navigating to About page: {about_url}")
            await page.wait_for_timeout(1000)
            await page.goto(about_url, wait_until="domcontentloaded", timeout=25000)
            await page.wait_for_timeout(2000)
            await human_mouse_move(page)
            await human_scroll(page, distance_total=600)
            
            about_data = await extract_details_from_page(page, page_url)
            
            # Merge results, prioritize non-N/A parameters
            for key in ["phone", "email", "website", "category"]:
                if data[key] == "N/A" and about_data[key] != "N/A":
                    data[key] = about_data[key]
            if data["name"] == "N/A" and about_data["name"] != "N/A":
                data["name"] = about_data["name"]
                
        return data
    except Exception as e:
        print(f"      [-] Error loading page {page_url}: {e}")
        return None

# ==========================================
# SEARCH PANEL & SCRAPING ROUTINE
# ==========================================
async def scrape_facebook_pages(queries, max_leads_per_query, headed=False):
    """
    Executes search queries sequentially, scrolls feeds to compile target URLs,
    extracts business leads, and applies SQLite-based deduplication rules.
    """
    init_db()
    results = []
    
    async with async_playwright() as p:
        context, is_new_auth = await setup_auth_context(p, headed)
        
        search_page = await context.new_page()
        
        for q_idx, query in enumerate(queries, 1):
            print(f"\n[*] [{q_idx}/{len(queries)}] Starting search for: '{query}'")
            try:
                search_url = f"https://www.facebook.com/search/pages/?q={urllib.parse.quote(query)}"
                print(f"    [*] Navigating to search feed: {search_url}")
                await search_page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
                await search_page.wait_for_timeout(3000)
                
                # Check redirect
                if "login" in search_page.url or "checkpoint" in search_page.url:
                    print("    [Warning] Facebook redirected to login page. Session expired.")
                    if is_new_auth:
                        print("    [-] Auth just completed but still redirected. Session is restricted.")
                    break
                
                page_urls = set()
                no_new_results_count = 0
                previous_count = 0
                
                # Scroll search result panel to discover listings
                scroll_attempts = 12
                for attempt in range(1, scroll_attempts + 1):
                    links = await search_page.locator('a').all()
                    for link in links:
                        try:
                            href = await link.get_attribute("href")
                            cleaned = clean_facebook_url(href)
                            if cleaned:
                                page_urls.add(cleaned)
                        except Exception:
                            pass
                            
                    current_count = len(page_urls)
                    print(f"      [Scroll Attempt {attempt}/{scroll_attempts}] Found {current_count} potential page links...")
                    
                    if current_count >= max_leads_per_query:
                        break
                    if current_count == previous_count:
                        no_new_results_count += 1
                        if no_new_results_count >= 4:
                            break
                    else:
                        no_new_results_count = 0
                        
                    previous_count = current_count
                    
                    # Human scrolling
                    await human_scroll(search_page, distance_total=600)
                    await search_page.wait_for_timeout(1500)
                
                # Deduplicate list and filter through SQLite history
                raw_listings = list(page_urls)[:max_leads_per_query]
                listings_to_scrape = []
                for url in raw_listings:
                    if is_duplicate(url, None):
                        # Skip duplicate page URLs
                        continue
                    listings_to_scrape.append(url)
                    
                print(f"    [*] Found {len(listings_to_scrape)} new leads to process sequentially...")
                
                # Extract page details sequentially (minimizes risk of session flag & block)
                for idx, page_url in enumerate(listings_to_scrape, 1):
                    # Check database again
                    if is_duplicate(page_url, None):
                        continue
                        
                    data = await scrape_single_page(search_page, page_url)
                    
                    if data:
                        # Extract and check phone duplicates
                        if data["phone"] != "N/A" and is_duplicate(None, data["phone"]):
                            print(f"      [{idx}/{len(listings_to_scrape)}] Skipping (Duplicate Phone: {data['phone']})")
                            # Human delay before next page
                            await random_delay()
                            continue
                            
                        # Success log
                        print(f"      [✔] Extracted: {data['name']} | Category: {data['category']} | Phone: {data['phone']} | Email: {data['email']} | Web: {data['website']}")
                        
                        is_custom = data["website"] != "N/A" and len(data["website"]) > 0
                        
                        lead_data = {
                            "Query": query,
                            "Business Name (اسم العمل)": data["name"],
                            "Business Category (التصنيف)": data["category"],
                            "Phone Number (رقم الهاتف)": data["phone"],
                            "Email (البريد الإلكتروني)": data["email"],
                            "Website URL (رابط الموقع)": data["website"],
                            "Is Custom Website (موقع مخصص)": "Yes" if is_custom else "No",
                            "Facebook Link (رابط فيسبوك)": page_url
                        }
                        
                        save_lead_to_db(lead_data)
                        results.append(lead_data)
                        
                    # Sequential delay between pages
                    await random_delay()
                    
                # Small delay between query searches
                await asyncio.sleep(random.uniform(3.0, 6.0))
                
            except Exception as query_err:
                print(f"[-] Error processing query '{query}': {query_err}")
                
        await context.close()
        
    return results

# ==========================================
# OUTPUT GENERATOR (STYLED EXCEL WORKBOOK)
# ==========================================
def format_excel_output(df, file_path):
    """
    Saves and formats data into a professionally styled Excel workbook.
    Highlights hot prospects (no custom website) in a soft yellow color fill.
    If file exists, appends data safely and applies styling over the entire dataset.
    """
    if os.path.exists(file_path):
        try:
            existing_df = pd.read_excel(file_path)
            combined_df = pd.concat([existing_df, df], ignore_index=True)
            combined_df.drop_duplicates(subset=["Facebook Link (رابط فيسبوك)"], keep="first", inplace=True)
            df = combined_df
            print(f"[*] Appending to existing Excel file. Total records in sheet: {len(df)}")
        except Exception as e:
            print(f"[-] Warning: Failed to read existing Excel file: {e}. Rewriting file.")

    import time
    writer = None
    max_retries = 5
    for attempt in range(1, max_retries + 1):
        try:
            writer = pd.ExcelWriter(file_path, engine="openpyxl")
            break
        except PermissionError:
            if attempt == max_retries:
                timestamp = datetime.now().strftime("%H%M%S")
                base, ext = os.path.splitext(file_path)
                file_path = f"{base}_{timestamp}{ext}"
                print(f"[Warning] Excel file is permanently locked. Saving output to alternative file path: {file_path}")
                writer = pd.ExcelWriter(file_path, engine="openpyxl")
            else:
                print(f"[Warning] Excel file is locked. Retrying in 2 seconds... (Attempt {attempt}/{max_retries})")
                time.sleep(2)
                
    df.to_excel(writer, sheet_name="Meta Leads", index=False)
    
    workbook = writer.book
    worksheet = writer.sheets["Meta Leads"]
    worksheet.views.sheetView[0].showGridLines = True
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Corporate Dark Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")  # Zebra stripe tint
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    flag_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft yellow highlight
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_idx, col in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    worksheet.row_dimensions[1].height = 28
    
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 22
        
        is_custom_web = worksheet.cell(row=row_idx, column=df.columns.get_loc("Is Custom Website (موقع مخصص)") + 1).value
        
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        if is_custom_web == "No":
            row_fill = flag_fill
            
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            
            col_name = df.columns[col_idx - 1]
            if any(term in col_name for term in ["Phone", "Email", "Website", "Link", "Is Custom"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val) + 4)
            else:
                max_len = max(max_len, len(val))
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    writer.close()
    print(f"\n[+] Lead output saved and formatted successfully in: {file_path}")

# ==========================================
# MAIN EXECUTION ROUTINE
# ==========================================
async def main():
    print("=" * 60)
    print("        ENTERPRISE META FACEBOOK PAGES LEAD SCRAPER")
    print("=" * 60)
    
    # Programmatically compile Query Matrix (Categories × Cities)
    queries = [f"{cat} في {city}" for cat in CATEGORIES for city in CITIES]
    
    # Generate timestamped filename
    today_str = datetime.now().strftime("%Y_%m_%d")
    output_filename = f"meta_leads_output_{today_str}.xlsx"
    output_path = os.path.join(SCRIPT_DIR, output_filename)
    
    print(f"[*] Dynamically generated {len(queries)} query combinations.")
    print(f"[*] Output destination: {output_path}")
    print(f"[*] Deduplication History: {DB_NAME}")
    print("=" * 60)
    
    # Default target listings per query
    max_leads_per_query = 20
    
    print(f"[*] Target max leads per query: {max_leads_per_query}")
    print("-" * 60)
    
    # Run the scraper
    results = []
    try:
        results = await scrape_facebook_pages(queries, max_leads_per_query, headed=False)
    except KeyboardInterrupt:
        print("\n[-] Process interrupted by user. Exporting whatever was collected...")
        
    if results:
        df = pd.DataFrame(results)
        format_excel_output(df, output_path)
    else:
        print("\n[-] No new leads were extracted during this run.")

if __name__ == "__main__":
    asyncio.run(main())
